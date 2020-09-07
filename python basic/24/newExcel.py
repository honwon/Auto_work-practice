from openpyxl import Workbook

# 클래스 생성
wb = Workbook()

# 지정한 이름의 시트 생성
ws = wb.create_sheet("sheet_test2")

# 지정 범위 내용 수정
ws["A1"] = "moheom"
ws["B2"] = "sibla"

# 행 단위로 데이터를 넣음
ws.append(["Number","Name"])

for i in range(10):
    ws.append([i, str(i) + " data"])

# 파일로 저장
wb.save("testTest2.xlsx")