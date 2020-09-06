from openpyxl import load_workbook

# read_only = 읽기 전용 엑세스 될 때만 정보 받아옴 행만 받아온다.
wb = load_workbook("big.xlsx",read_only=True)
data = wb.active

# iter_rows 행단위로 묶어줌 행 열 개수 지정 가능
for row in data.iter_rows(max_col=1,max_row=1):
    for cell in row:
        print(cell.value)